
import io
import requests
from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment


LOGO_URL = "https://media.licdn.com/dms/image/v2/C4D0BAQFynSl_Yj90cQ/company-logo_200_200/company-logo_200_200/0/1630472942468/vicente_monteiro_advogados_logo?e=2147483647&v=beta&t=HUb5xhVbshv-LGKYpmpkkuJfUxX30S5oMjefFv7jM4s"
BACKGROUND_URL = "https://raw.githubusercontent.com/leomarjms26-netizen/planilha-clientes/refs/heads/main/vm-4.jpg"
VALOR_HORA_DEFAULT = 462.62

st.markdown(
    f"""
    <style>
    html, body, [class*="stAppViewContainer"], [class*="stApp"], [data-testid="stAppViewContainer"] {{
        background: 
            linear-gradient(rgba(5,85,119,0.75), rgba(5,85,119,0.75)),
            url('{BACKGROUND_URL}') !important;
        background-size: cover !important;
        background-position: center center !important;
        background-attachment: fixed !important;
    }}

    [data-testid="stHeader"], [data-testid="stToolbar"], [data-testid="stSidebar"] {{
        background: rgba(5,85,119,0.7) !important;
        backdrop-filter: blur(6px);
    }}

    /* Texto e tÃ­tulos em branco */
    h1, h2, h3, h4, h5, h6, p, label, span, div {{
        color: #f8f9fa !important;
    }}

    /* Caixas escuras translÃºcidas para componentes */
    .stFileUploader, .stDownloadButton, .stTextInput, .stSelectbox, .stDataFrame, .stAlert {{
        background-color: rgba(0, 0, 0, 0.55) !important;
        border-radius: 12px;
        padding: 14px;
        color: #ffffff !important;
    }}

    /* Borda e sombra sutil */
    .stFileUploader, .stDownloadButton {{
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.4);
        border: 1px solid rgba(255,255,255,0.1);
    }}

    /* BotÃµes */
    button[kind="primary"], .stDownloadButton > button {{
        background-color: rgba(5,85,119,0.9) !important;
        color: #ffffff !important;
        border: none !important;
    }}
    button[kind="primary"]:hover, .stDownloadButton > button:hover {{
        background-color: rgba(5,100,140,1) !important;
    }}
    </style>
    """,
    unsafe_allow_html=True
)

st.set_page_config(page_title="Cliente Mensais", layout="wide")
st.title("âš–ï¸ Clientes Mensais")

uploaded_file = st.file_uploader("ðŸ“¤ Envie a planilha BRUTA (xlsx/xls)", type=["xlsx", "xls"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    st.write(f"âœ… Planilha carregada com {len(df)} linhas.")

    coluna1, coluna2 = df.columns[0], df.columns[1]
    df["CLIENTES"] = df[[coluna1, coluna2]].fillna("").agg(" ".join, axis=1).str.strip()

    colunas_para_manter = [
        "CLIENTES",
        "DuraÃ§Ã£o",
        "Data de inÃ­cio",
        "Executante",
        "DescriÃ§Ã£o",
        "VÃ­nculos com processo / NÃºmero de CNJ",
        "ContrÃ¡rio principal / Nome/razÃ£o social",
        "VÃ­nculos com processo / Pasta"
    ]
    df_final = df[[col for col in colunas_para_manter if col in df.columns]].copy()

    if "DuraÃ§Ã£o" in df_final.columns:
        df_final["DuraÃ§Ã£o"] = df_final["DuraÃ§Ã£o"].apply(lambda x: str(x).split()[-1])
    if "Data de inÃ­cio" in df_final.columns:
        df_final["Data de inÃ­cio"] = pd.to_datetime(df_final["Data de inÃ­cio"], errors="coerce").dt.strftime("%Y-%m-%d")

    def somar_duracoes(series):
        tempos = pd.to_timedelta(series, errors='coerce')
        total_segundos = tempos.dt.total_seconds().sum()
        horas = int(total_segundos // 3600)
        minutos = int((total_segundos % 3600) // 60)
        segundos = int(total_segundos % 60)
        return f"{horas:02}:{minutos:02}:{segundos:02}", total_segundos / 3600  # HH:MM:SS e decimal

    response = requests.get(LOGO_URL)

    wb = Workbook()
    wb.remove(wb.active)
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)

    contador_tabela = 1 

    def criar_aba(nome, df_dados, valor_hora):
        global contador_tabela
        ws = wb.create_sheet(title=nome[:31])

        logo_stream = BytesIO(response.content)
        img = Image(logo_stream)
        img.width = 120
        img.height = 120
        ws.add_image(img, "A1")

        start_row_horas = 8
        horas_totais_str, horas_totais_decimal = somar_duracoes(df_dados["DuraÃ§Ã£o"])
        ws[f"A{start_row_horas}"] = "HORAS TOTAIS"
        ws[f"B{start_row_horas}"] = horas_totais_str
        ws[f"A{start_row_horas+1}"] = "VALOR HORA"
        ws[f"B{start_row_horas+1}"] = valor_hora
        ws[f"A{start_row_horas+2}"] = "TOTAL MENSAL"
        ws[f"B{start_row_horas+2}"] = f"=B{start_row_horas+3}*B{start_row_horas+1}"
        ws[f"B{start_row_horas+3}"] = horas_totais_decimal
        ws.row_dimensions[start_row_horas+3].hidden = True

        for row in range(start_row_horas, start_row_horas + 3):
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = header_align
                if col == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        start_row_tabela = start_row_horas + 5
        for r_idx, row in enumerate(dataframe_to_rows(df_dados, index=False, header=True), start=start_row_tabela):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = cell_align
                if r_idx == start_row_tabela:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align

        max_row = ws.max_row
        max_col = ws.max_column
        last_col_letter = get_column_letter(max_col)
        table_ref = f"A{start_row_tabela}:{last_col_letter}{max_row}"
        tabela_nome = f"TABELA_{contador_tabela}"
        contador_tabela += 1
        tab = Table(displayName=tabela_nome, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

    criar_aba("GERAL", df_final, VALOR_HORA_DEFAULT)

    for cliente in df_final["CLIENTES"].unique():
        df_cliente = df_final[df_final["CLIENTES"] == cliente]
        criar_aba(str(cliente), df_cliente, VALOR_HORA_DEFAULT)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    st.success(f"âœ… Planilha gerada com {len(df_final['CLIENTES'].unique())+1} abas (GERAL + clientes).")
    st.download_button(
        "â¤“ Baixar planilha final",
        data=bio,
        file_name="Planilha_Final_Clientes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )





